import logging
from typing import List, Optional
import uuid
import yaml

from fastapi import APIRouter, Depends, HTTPException, status, BackgroundTasks
from sqlalchemy.orm import Session
from datetime import datetime, timezone

from app.db.session import get_db
from app.db.data_contract import DataContractModel
from app.db.data_asset import DataAssetModel
from pydantic import BaseModel
from app.api.deps import get_current_user

router = APIRouter()
logger = logging.getLogger(__name__)

class DataContractResponse(BaseModel):
    id: str
    dataset_id: str
    yaml_content: str
    version: int
    is_active: bool
    created_at: datetime
    created_by: Optional[str]
    
    # Asset fields
    catalog: Optional[str] = None
    schema_name: Optional[str] = None
    table_name: Optional[str] = None
    data_quality: Optional[dict] = None
    certification_violations: Optional[List[str]] = None
    # Full per-rule checklist (pass + fail, with category) from the last policy
    # evaluation, so the UI can render the identical Sentinel checklist.
    certification_rule_results: Optional[List[dict]] = None
    certified: Optional[bool] = False
    last_synced_at: Optional[datetime] = None

    model_config = {
        "from_attributes": True,
        "populate_by_name": True
    }

class DataContractCreate(BaseModel):
    dataset_id: str
    yaml_content: str

@router.post("/sync")
async def sync_contracts(
    background_tasks: BackgroundTasks,
    force: bool = False,
    dataset_id: Optional[str] = None,
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    from app.tools.governance.draft_odcs import draft_odcs_contract
    from app.tools.execute_workflow import execute_workflow
    from app.providers.databricks import DatabricksProvider
    from app.core.config import settings
    
    try:
        # 1. Query Databricks for tables with the 'dataset' tag
        provider = DatabricksProvider(
            host=settings.DATABRICKS_HOST or settings.DATABRICKS_WORKSPACE_URL,
            token=settings.DATABRICKS_TOKEN,
            client_id=settings.DATABRICKS_CLIENT_ID,
            client_secret=settings.DATABRICKS_CLIENT_SECRET,
            config={"warehouse_id": settings.DATABRICKS_WAREHOUSE_ID}
        )
        
        dataset_groups = {}
        
        if dataset_id:
            # Sync a specific dataset
            parts = dataset_id.split(".")
            if len(parts) == 3:
                catalog_name, schema_name, table_name = parts
                query = f"SELECT catalog_name, schema_name, table_name, tag_value FROM {catalog_name}.information_schema.table_tags WHERE tag_name = 'dataset' AND catalog_name = '{catalog_name}' AND schema_name = '{schema_name}' AND table_name = '{table_name}'"
                try:
                    response = provider.client.statement_execution.execute_statement(
                        statement=query,
                        warehouse_id=settings.DATABRICKS_WAREHOUSE_ID,
                        wait_timeout="30s"
                    )
                    if response.result and response.result.data_array:
                        for row in response.result.data_array:
                            catalog_name, schema_name, table_name, dataset_name = row
                            if not dataset_name or not str(dataset_name).strip():
                                logger.warning(
                                    f"Skipping table {catalog_name}.{schema_name}.{table_name}: "
                                    f"'dataset' tag has an empty value"
                                )
                                continue
                            full_name = f"{catalog_name}.{schema_name}.{table_name}"
                            if dataset_name not in dataset_groups:
                                dataset_groups[dataset_name] = []
                            dataset_groups[dataset_name].append(full_name)
                    else:
                        # If not found in table_tags, maybe it's the dataset_name itself
                        pass
                except Exception as e:
                    logger.warning(f"Could not query information_schema for {dataset_id}: {e}")
            
            # If we didn't find it by table name, maybe dataset_id is the tag value
            if not dataset_groups:
                catalogs = provider.client.catalogs.list()
                for catalog in catalogs:
                    if catalog.name in ("system", "samples"):
                        continue
                    query = f"SELECT catalog_name, schema_name, table_name, tag_value FROM {catalog.name}.information_schema.table_tags WHERE tag_name = 'dataset' AND tag_value = '{dataset_id}'"
                    try:
                        response = provider.client.statement_execution.execute_statement(
                            statement=query,
                            warehouse_id=settings.DATABRICKS_WAREHOUSE_ID,
                            wait_timeout="30s"
                        )
                        if response.result and response.result.data_array:
                            for row in response.result.data_array:
                                catalog_name, schema_name, table_name, dataset_name = row
                                if not dataset_name or not str(dataset_name).strip():
                                    logger.warning(
                                        f"Skipping table {catalog_name}.{schema_name}.{table_name}: "
                                        f"'dataset' tag has an empty value"
                                    )
                                    continue
                                full_name = f"{catalog_name}.{schema_name}.{table_name}"
                                if dataset_name not in dataset_groups:
                                    dataset_groups[dataset_name] = []
                                dataset_groups[dataset_name].append(full_name)
                    except Exception as e:
                        logger.warning(f"Could not query information_schema for catalog {catalog.name}: {e}")
        else:
            # Fetch all catalogs the SP has access to
            catalogs = provider.client.catalogs.list()
            
            for catalog in catalogs:
                if catalog.name in ("system", "samples"):
                    continue
                    
                # Query the local information_schema for each catalog
                query = f"SELECT catalog_name, schema_name, table_name, tag_value FROM {catalog.name}.information_schema.table_tags WHERE tag_name = 'dataset'"
                logger.info(f"Querying information_schema for catalog {catalog.name}")
                
                try:
                    response = provider.client.statement_execution.execute_statement(
                        statement=query,
                        warehouse_id=settings.DATABRICKS_WAREHOUSE_ID,
                        wait_timeout="30s"
                    )
                    
                    if response.result and response.result.data_array:
                        logger.info(f"Found {len(response.result.data_array)} tagged tables in catalog {catalog.name}")
                        for row in response.result.data_array:
                            catalog_name, schema_name, table_name, dataset_name = row
                            if not dataset_name or not str(dataset_name).strip():
                                logger.warning(
                                    f"Skipping table {catalog_name}.{schema_name}.{table_name}: "
                                    f"'dataset' tag has an empty value"
                                )
                                continue
                            full_name = f"{catalog_name}.{schema_name}.{table_name}"
                            if dataset_name not in dataset_groups:
                                dataset_groups[dataset_name] = []
                            dataset_groups[dataset_name].append(full_name)
                    else:
                        logger.debug(f"No tables found with 'dataset' tag in catalog {catalog.name}")
                except Exception as e:
                    logger.warning(f"Could not query information_schema for catalog {catalog.name}: {e}")
                    
        logger.info(f"Discovery complete. Found {len(dataset_groups)} unique data sets.")
        
        if not dataset_groups:
            if dataset_id:
                # Still run background to clean up if it was deleted
                background_tasks.add_task(run_sync_contracts_background, dataset_groups, force, dataset_id)
            return {"status": "success", "message": "No tables found with 'dataset' tag."}

        background_tasks.add_task(run_sync_contracts_background, dataset_groups, force, dataset_id)
        return {"status": "success", "message": f"Sync started in the background for {len(dataset_groups)} data sets. This may take a few minutes."}
        
    except Exception as e:
        logger.error(f"Failed to sync contracts: {e}")
        raise HTTPException(status_code=500, detail=str(e))

async def run_sync_contracts_background(dataset_groups: dict, force: bool, specific_dataset_id: Optional[str] = None):
    from app.db.session import get_lakebase_session
    import hashlib
    import json
    from app.tools.governance.draft_odcs import fetch_datasets_metadata
    from app.tools.governance.draft_odcs import draft_odcs_contract
    
    db = get_lakebase_session()
    try:
        requests_created = 0
        
        # Clean up contracts for datasets that are no longer tagged
        if specific_dataset_id:
            if specific_dataset_id not in dataset_groups:
                logger.info(f"Dataset {specific_dataset_id} is no longer tagged. Deleting contract.")
                db.query(DataContractModel).filter(DataContractModel.dataset_id == specific_dataset_id).delete()
                asset = db.query(DataAssetModel).filter(DataAssetModel.id == specific_dataset_id).first()
                if asset:
                    asset.contract_url = None
                    asset.certified = False
                    db.add(asset)
        else:
            active_contracts = db.query(DataContractModel).filter(DataContractModel.is_active == True).all()
            for contract in active_contracts:
                if contract.dataset_id not in dataset_groups:
                    logger.info(f"Dataset {contract.dataset_id} is no longer tagged. Deleting contract.")
                    db.query(DataContractModel).filter(DataContractModel.dataset_id == contract.dataset_id).delete()
                    asset = db.query(DataAssetModel).filter(DataAssetModel.id == contract.dataset_id).first()
                    if asset:
                        asset.contract_url = None
                        asset.certified = False
                        db.add(asset)
        db.commit()

        for dataset_name, table_ids in dataset_groups.items():
            # Check if we have an existing contract
            existing_contract = db.query(DataContractModel).filter(
                DataContractModel.dataset_id == dataset_name,
                DataContractModel.is_active == True
            ).first()
            
            # Fetch current metadata for the tables
            try:
                datasets_metadata = await fetch_datasets_metadata(table_ids)
            except Exception as e:
                logger.error(f"Failed to fetch metadata for {dataset_name}: {e}")
                continue
            
            # Update the DataAsset catalog text (e.g. "3 tables, 1 view") unconditionally
            asset = db.query(DataAssetModel).filter(DataAssetModel.id == dataset_name).first()
            
            tables_count = sum(1 for m in datasets_metadata if "VIEW" not in str(m.get("table_type", "")).upper())
            views_count = sum(1 for m in datasets_metadata if "VIEW" in str(m.get("table_type", "")).upper())
            
            parts = []
            if tables_count > 0:
                parts.append(f"{tables_count} table{'s' if tables_count > 1 else ''}")
            if views_count > 0:
                parts.append(f"{views_count} view{'s' if views_count > 1 else ''}")
            
            summary_str = ", ".join(parts) if parts else "Multiple datasets"
            
            if asset:
                asset.contract_url = f"/governance/certification?dataset={dataset_name}"
                # Refresh the count summary whenever we actually resolved real
                # tables/views. This corrects placeholder subtitles (e.g. a
                # sentinel-seeded "Multiple datasets") that would otherwise stick
                # forever. When no tables could be counted we leave the existing
                # value alone rather than clobbering a good count with a placeholder.
                if parts:
                    asset.catalog = summary_str
                    asset.schema = ""
                db.add(asset)
            else:
                asset = DataAssetModel(
                    id=dataset_name,
                    catalog=summary_str,
                    schema="",
                    table_name=dataset_name,
                    type="DATA_PRODUCT",
                    description=f"Data contract for {dataset_name}",
                    domain="unknown",
                    contract_url=f"/governance/certification?dataset={dataset_name}"
                )
                db.add(asset)
            db.commit()
            
            # Calculate metadata hash
            metadata_hash = hashlib.md5(json.dumps(datasets_metadata, sort_keys=True).encode()).hexdigest()
            
            # Check if tables in YAML match tables in metadata
            yaml_tables_match = True
            if existing_contract and existing_contract.yaml_content:
                try:
                    parsed_yaml = yaml.safe_load(existing_contract.yaml_content)
                    if "schema" in parsed_yaml and isinstance(parsed_yaml["schema"], list):
                        yaml_table_names = set()
                        for table_def in parsed_yaml["schema"]:
                            phys_name = table_def.get("physicalName", "")
                            if phys_name:
                                yaml_table_names.add(phys_name)
                        
                        # Use datasets_metadata instead of table_ids because datasets_metadata 
                        # only contains tables that actually exist and could be fetched
                        meta_table_names = {m["dataset_id"] for m in datasets_metadata}
                        meta_short_names = {m["table"] for m in datasets_metadata}
                        
                        # Check if any table in YAML is missing from metadata
                        for y_name in yaml_table_names:
                            if y_name not in meta_table_names and y_name not in meta_short_names:
                                logger.info(f"Table {y_name} in YAML is missing from metadata. Forcing LLM call.")
                                yaml_tables_match = False
                                break
                                
                        # Also check if any table in metadata is missing from YAML
                        if yaml_tables_match:
                            for m_name in meta_table_names:
                                short_m_name = m_name.split('.')[-1]
                                if m_name not in yaml_table_names and short_m_name not in yaml_table_names:
                                    logger.info(f"Table {m_name} in metadata is missing from YAML. Forcing LLM call.")
                                    yaml_tables_match = False
                                    break
                except Exception as e:
                    logger.warning(f"Failed to parse existing YAML to check tables: {e}")
                    yaml_tables_match = False
            
            if existing_contract and hasattr(existing_contract, 'metadata_hash') and existing_contract.metadata_hash == metadata_hash and yaml_tables_match and not force:
                logger.info(f"No metadata changes detected for {dataset_name}, skipping LLM call.")
                continue
            
            existing_yaml = existing_contract.yaml_content if existing_contract else None
            
            # Draft or update the ODCS
            try:
                logger.info(f"Starting LLM ODCS drafting for {dataset_name} (tables: {table_ids})...")
                odcs_yaml = await draft_odcs_contract._func(
                    dataset_ids=table_ids, 
                    existing_odcs_yaml=existing_yaml,
                    pre_fetched_metadata=datasets_metadata
                )
                logger.info(f"Completed LLM ODCS drafting for {dataset_name}. Received {len(odcs_yaml) if odcs_yaml else 0} chars.")
            except Exception as e:
                logger.error(f"Error drafting ODCS for dataset {dataset_name}: {e}")
                continue
                
            if not odcs_yaml or "apiVersion:" not in odcs_yaml or "kind: DataContract" not in odcs_yaml:
                logger.error(f"LLM returned an invalid ODCS format for {dataset_name}. It may have timed out or hit rate limits.")
                continue
            
            # Check if the contract actually changed
            if existing_yaml and existing_yaml.strip() == odcs_yaml.strip():
                logger.info(f"No YAML changes detected for contract {dataset_name}, skipping update.")
                if hasattr(existing_contract, 'metadata_hash'):
                    existing_contract.metadata_hash = metadata_hash
                    db.commit()
                continue
            
            # Save directly to database
            new_version = (existing_contract.version + 1) if existing_contract else 1
            if existing_contract:
                existing_contract.is_active = False
                db.add(existing_contract)
                
            new_contract = DataContractModel(
                id=str(uuid.uuid4()),
                dataset_id=dataset_name,
                yaml_content=odcs_yaml,
                version=new_version,
                is_active=True,
                created_at=datetime.now(timezone.utc),
                metadata_hash=metadata_hash
            )
            db.add(new_contract)
            
            db.commit()
            requests_created += 1

        logger.info(f"Background sync complete. Synced {len(dataset_groups)} data sets. Updated {requests_created} data contracts.")
        
    except Exception as e:
        logger.error(f"Background sync failed: {e}")
    finally:
        db.close()

@router.get("", response_model=List[DataContractResponse])
@router.get("/", response_model=List[DataContractResponse])
def list_contracts(db: Session = Depends(get_db)):
    """List all active data contracts."""
    import json
    contracts = db.query(DataContractModel).filter(DataContractModel.is_active == True).all()
    results = []
    for c in contracts:
        asset = db.query(DataAssetModel).filter(DataAssetModel.id == c.dataset_id).first()
        contract_dict = {
            "id": c.id,
            "dataset_id": c.dataset_id,
            "yaml_content": c.yaml_content,
            "version": c.version,
            "is_active": c.is_active,
            "created_at": c.created_at,
            "created_by": c.created_by,
        }
        if asset:
            violations = asset.certification_violations
            if isinstance(violations, str):
                try:
                    violations = json.loads(violations)
                except Exception:
                    violations = None

            rule_results = asset.certification_rule_results
            if isinstance(rule_results, str):
                try:
                    rule_results = json.loads(rule_results)
                except Exception:
                    rule_results = None

            contract_dict.update({
                "catalog": asset.catalog,
                "schema_name": asset.schema,
                "table_name": asset.table_name,
                "data_quality": asset.data_quality,
                "certification_violations": violations,
                "certification_rule_results": rule_results,
                "certified": asset.certified,
                "last_synced_at": asset.last_synced_at
            })
        results.append(contract_dict)
    return results


# Category display order for the exec report. Mirrors the buckets defined in
# ``policies/data_certification.rego`` (rule_category); "Other" catches any rule
# not yet mapped so a new rego rule never silently drops out of the report.
_REPORT_CATEGORY_ORDER = ["Structure", "Metadata", "Tagging", "Data Quality"]


def _category_status(rule_results: list, category: str) -> str:
    """Roll a category's rules up to a single pass/fail/n-a for the exec sheet."""
    rules = [r for r in (rule_results or []) if (r.get("category") or "Other") == category]
    if not rules:
        return "n/a"
    return "fail" if any(not r.get("passed") for r in rules) else "pass"


@router.get("/certification-report")
def certification_report(db: Session = Depends(get_db)):
    """Download an XLSX certification report for leadership.

    Sheet 1 (Overview): one row per dataset with a green ``pass`` / red ``fail``
    cell per high-level category (Structure, Metadata, Tagging, Data Quality).
    Sheet 2 (Details): one row per exact failure (dataset x failed check x
    message) for the teams that need to remediate.
    """
    import json
    import re
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from app.core.config import settings

    environment = (settings.ENVIRONMENT or "unknown").strip()
    generated_at = datetime.now(timezone.utc)

    contracts = (
        db.query(DataContractModel)
        .filter(DataContractModel.is_active == True)  # noqa: E712
        .all()
    )
    assets = {a.id: a for a in db.query(DataAssetModel).all()}

    def _rule_results(asset) -> list:
        rr = getattr(asset, "certification_rule_results", None) if asset else None
        if isinstance(rr, str):
            try:
                rr = json.loads(rr)
            except Exception:
                rr = None
        return rr or []

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="374151", end_color="374151", fill_type="solid")
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    pass_font = Font(color="006100", bold=True)
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    fail_font = Font(color="9C0006", bold=True)
    na_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    na_font = Font(color="9CA3AF")
    center = Alignment(horizontal="center", vertical="center")

    wb = Workbook()

    # --- Sheet 1: Overview (exec) ---
    ws = wb.active
    ws.title = "Overview"
    overview_headers = ["Dataset", "Status"] + _REPORT_CATEGORY_ORDER
    last_col = len(overview_headers)

    # Metadata block (environment + generation time) above the table. Written at
    # fixed rows so the table's header/freeze/filter positions stay deterministic
    # (row 4 is left blank as a spacer; the table header sits on row 5).
    title_font = Font(bold=True, size=14)
    meta_font = Font(bold=True, color="374151")
    ws.cell(row=1, column=1, value="Data Certification Report").font = title_font
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    ws.cell(row=2, column=1, value="Environment").font = meta_font
    ws.cell(row=2, column=2, value=environment)
    ws.cell(row=3, column=1, value="Generated (UTC)").font = meta_font
    ws.cell(row=3, column=2, value=generated_at.strftime("%Y-%m-%d %H:%M:%S"))

    header_row = 5
    for col, header in enumerate(overview_headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    detail_rows = []  # collected for sheet 2

    sorted_contracts = sorted(contracts, key=lambda c: (c.dataset_id or "").lower())
    for c in sorted_contracts:
        asset = assets.get(c.dataset_id)
        rule_results = _rule_results(asset)
        name = (asset.table_name if asset and asset.table_name else c.dataset_id)
        status = "Certified" if (asset and asset.certified) else "Uncertified"
        row = [name, status]
        for cat in _REPORT_CATEGORY_ORDER:
            row.append(_category_status(rule_results, cat))
        ws.append(row)
        r = ws.max_row
        status_cell = ws.cell(row=r, column=2)
        status_cell.alignment = center
        if status == "Certified":
            status_cell.font = pass_font
        for i, cat in enumerate(_REPORT_CATEGORY_ORDER):
            cell = ws.cell(row=r, column=3 + i)
            cell.alignment = center
            val = cell.value
            if val == "pass":
                cell.fill, cell.font = pass_fill, pass_font
            elif val == "fail":
                cell.fill, cell.font = fail_fill, fail_font
            else:
                cell.value = "n/a"
                cell.fill, cell.font = na_fill, na_font

        # Collect exact failures for the detail sheet.
        for rr in rule_results:
            if rr.get("passed"):
                continue
            msgs = rr.get("messages") or []
            check = rr.get("description") or rr.get("id") or ""
            category = rr.get("category") or "Other"
            if msgs:
                for m in msgs:
                    detail_rows.append([name, category, check, m])
            else:
                detail_rows.append([name, category, check, "Failed"])

    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(overview_headers))}{ws.max_row}"
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 14
    for i in range(len(_REPORT_CATEGORY_ORDER)):
        ws.column_dimensions[get_column_letter(3 + i)].width = 16

    # --- Sheet 2: Details ---
    ws2 = wb.create_sheet("Failure Details")
    detail_headers = ["Dataset", "Category", "Failed Check", "Detail"]
    ws2.append(detail_headers)
    for col in range(1, len(detail_headers) + 1):
        cell = ws2.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
    if detail_rows:
        for dr in detail_rows:
            ws2.append(dr)
    else:
        ws2.append(["All datasets pass every check.", "", "", ""])
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(detail_headers))}{ws2.max_row}"
    ws2.column_dimensions["A"].width = 34
    ws2.column_dimensions["B"].width = 16
    ws2.column_dimensions["C"].width = 44
    ws2.column_dimensions["D"].width = 80

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    stamp = generated_at.strftime("%Y%m%d")
    env_slug = re.sub(r"[^A-Za-z0-9._-]+", "-", environment).strip("-").lower() or "unknown"
    filename = f"data-certification-report-{env_slug}-{stamp}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/{dataset_id}", response_model=List[DataContractResponse])
def get_contract_history(dataset_id: str, db: Session = Depends(get_db)):
    """Get the version history for a specific dataset contract."""
    import json
    contracts = db.query(DataContractModel).filter(
        DataContractModel.dataset_id == dataset_id
    ).order_by(DataContractModel.version.desc()).all()
    
    asset = db.query(DataAssetModel).filter(DataAssetModel.id == dataset_id).first()
    results = []
    for c in contracts:
        contract_dict = {
            "id": c.id,
            "dataset_id": c.dataset_id,
            "yaml_content": c.yaml_content,
            "version": c.version,
            "is_active": c.is_active,
            "created_at": c.created_at,
            "created_by": c.created_by,
        }
        if asset:
            violations = asset.certification_violations
            if isinstance(violations, str):
                try:
                    violations = json.loads(violations)
                except Exception:
                    violations = None

            rule_results = asset.certification_rule_results
            if isinstance(rule_results, str):
                try:
                    rule_results = json.loads(rule_results)
                except Exception:
                    rule_results = None

            contract_dict.update({
                "catalog": asset.catalog,
                "schema_name": asset.schema,
                "table_name": asset.table_name,
                "data_quality": asset.data_quality,
                "certification_violations": violations,
                "certification_rule_results": rule_results,
                "certified": asset.certified,
                "last_synced_at": asset.last_synced_at
            })
        results.append(contract_dict)
    return results

@router.post("", response_model=DataContractResponse)
@router.post("/", response_model=DataContractResponse)
def create_contract(contract: DataContractCreate, db: Session = Depends(get_db)):
    """Create a new version of a data contract."""
    # Validate YAML
    try:
        yaml.safe_load(contract.yaml_content)
    except yaml.YAMLError as e:
        raise HTTPException(status_code=400, detail=f"Invalid YAML: {str(e)}")

    # Find the latest version
    latest = db.query(DataContractModel).filter(
        DataContractModel.dataset_id == contract.dataset_id
    ).order_by(DataContractModel.version.desc()).first()

    new_version = (latest.version + 1) if latest else 1

    # Deactivate the old version
    if latest:
        latest.is_active = False
        db.add(latest)

    new_contract = DataContractModel(
        id=str(uuid.uuid4()),
        dataset_id=contract.dataset_id,
        yaml_content=contract.yaml_content,
        version=new_version,
        is_active=True,
        created_at=datetime.now(timezone.utc)
        # created_by could be pulled from auth context if needed
    )

    db.add(new_contract)
    
    # Also update the DataAsset to mark it as having a contract
    asset = db.query(DataAssetModel).filter(DataAssetModel.id == contract.dataset_id).first()
    if asset:
        asset.contract_url = f"/governance/certification?dataset={contract.dataset_id}"
        db.add(asset)
    else:
        # Create a mock asset so it shows up in the UI right away
        parts = contract.dataset_id.split(".")
        catalog = parts[0]
        schema = parts[1] if len(parts) > 1 else "default"
        table = parts[2] if len(parts) > 2 else "unknown"
        
        # Parse yaml to get some info
        try:
            parsed = yaml.safe_load(contract.yaml_content)
            description = parsed.get("description", {}).get("purpose", f"Data contract for {contract.dataset_id}")
            domain = parsed.get("domain", "unknown")
        except:
            description = f"Data contract for {contract.dataset_id}"
            domain = "unknown"

        asset = DataAssetModel(
            id=contract.dataset_id,
            catalog=catalog,
            schema=schema,
            table_name=table,
            type="TABLE",
            description=description,
            domain=domain,
            contract_url=f"/governance/certification?dataset={contract.dataset_id}"
        )
        db.add(asset)

    db.commit()
    db.refresh(new_contract)
    return new_contract

@router.post("/{dataset_id}/check-policy")
async def check_policy(
    dataset_id: str,
    background_tasks: BackgroundTasks,
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Run a policy check for a specific dataset."""
    from app.db.request import RequestModel
    from app.models.request import RequestType
    
    try:
        # Create an enforcement_sentinel request
        # The poller will pick this up and run the state machine
        request_id = f"req-{uuid.uuid4().hex[:8]}"
        
        new_request = RequestModel(
            id=request_id,
            type=RequestType.ENFORCEMENT_SENTINEL.value,
            title=f"Manual Policy Check for {dataset_id}",
            status="pending",
            environment="production",
            requester_email=current_user.email,
            created_at=datetime.now(timezone.utc),
            updated_at=datetime.now(timezone.utc),
            state_context={
                "dataset_id": dataset_id,
                "policies": ["data_certification"],
            },
        )
        
        db.add(new_request)
        db.commit()
        
        return {"status": "success", "message": f"Policy check started for {dataset_id}. Request ID: {request_id}"}
    except Exception as e:
        logger.error(f"Failed to check policy for {dataset_id}: {e}")
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

@router.delete("/{dataset_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_contract(
    dataset_id: str, 
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Delete all versions of a data contract and unset the asset contract_url."""
    # Check permissions - usually only admins or owners should delete contracts
    if not current_user.has_role("platform_admin") and not current_user.has_role("governance_admin"):
        raise HTTPException(status_code=403, detail="Not authorized to delete data contracts")
        
    contracts = db.query(DataContractModel).filter(DataContractModel.dataset_id == dataset_id).all()
    asset = db.query(DataAssetModel).filter(DataAssetModel.id == dataset_id).first()
    
    if not contracts and not asset:
        raise HTTPException(status_code=404, detail="Data contract not found")
        
    for contract in contracts:
        db.delete(contract)
        
    # Clear the contract_url from the DataAsset if it exists
    if asset:
        asset.contract_url = None
        asset.certified = False
        db.add(asset)
        
    db.commit()
    return None
